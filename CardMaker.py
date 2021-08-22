import getopt, sys, os
import openpyxl
from openpyxl import load_workbook
from math import ceil

#first serious personal project to learn some python and make my job easier
#mineral donations are stored in an excel sheet and would require manaual entry into a word document, upon seeing a
#donation which numbered in the thousands i was motivated to automate it.

#could benefit from a good amount of polish, havent touched it since early August 2021, consdier it depreciating?
#could potentially implement a system to create the destination sheet as it runs instead of using preset sized sheets.
#Destination sheets were originally in .docx and recreating them in excel made them a mess, works but its a pain

inputpath = input("Filename (Include .xlsx): ")
cardsize = int(input("Template Size (252 or 504): ")) #what size card file to use
#originally commandline input, changed to runtime input due to running on windows and intended users not being experienced in CS

if cardsize == 252 or cardsize == 504:
    pass
else:
    print("Invalid Template Size, Exiting...")
    exit()
startval = int(input("Starting Donation Value: "))
if startval < 1:
    print("Invalid Starting Value, Exiting...")
    exit()

#loading data sheet
src_sheet = load_workbook(filename=inputpath).active
wb = load_workbook(filename=f"CardTemplate{cardsize}.xlsx")
dest_sheet = wb.active
max_rows = src_sheet.max_row

mineral_list, locality_list, price_list = [], [], []    #creating empty lists
#lists are 0 indexed, data sheets are 2 indexed
count = 0

#depending on the format of the data sheet, first row value may change
for i in range(2, max_rows):
    don_num = src_sheet.cell(row = i, column = 1)
    if don_num.value == None:
        #checking to see if the end of the donations has been reached
        #there are less donations than max_rows in sheet
        break
    else:
        #loading source sheet into empty lists
        mineral = src_sheet.cell(row = i, column = 2).value
        locality = src_sheet.cell(row=i, column=3).value
        price = src_sheet.cell(row=i, column=4).value

        mineral_list.append(mineral.title())
        locality_list.append(locality.title())
        price_list.append(price)

page_num = int(cardsize / 18) #page_num is used to determine the next looping, each page holds 18 cards
#each of the destination sheets is 53 rows tall, bit of a magic number
#rest of the code is iterating through the destination sheet and inputting information
#cards are filled in left-to-right and then down, just like reading, goes card by card
for x in range(1,(page_num * 53),9):
    don_num = src_sheet.cell(row=(count + startval+1), column=1)
    if don_num.value == None:
        print("End of Donations Reached")
        print(f"Pages Used: {ceil(count / 18)}") #this is for printing and ease of use, tells you how much to print
        print(f"Excess Cards: {18 - ceil(count % 18)}")
        print(f"Last Donation Entered: {count}") #this is for if the donation is larger than the card templates
        break
    for i in range(3,14,5):
        don_num = src_sheet.cell(row=(count + startval + 1), column=1)
        if count > (cardsize - 1): #-1 because off by one error
            print("End of File Reached, Saving...")
            print(f"Pages Used: {ceil(count / 18)}")
            wb.save("CardOut.xlsx")
            os.startfile("CardOut.xlsx")
            exit()
        if don_num.value == None:
            break
        else:
            #(x,i) corresponds to the price cell
            #print(f"({x},{i})")
            #some magic numbers to get the
            dest_sheet.cell(row=x, column=i).value = f"${src_sheet.cell(row=(count + startval + 1), column=4).value}" #price cell
            dest_sheet.cell(row=(x+3), column=i).value = count + startval #donation number
            dest_sheet.cell(row=(x+4), column=(i-2)).value = f" Name: {src_sheet.cell(row=(count + startval + 1), column=2).value}" #name cell
            dest_sheet.cell(row=(x+6), column=(i-2)).value = f" Locality: {src_sheet.cell(row=(count + startval + 1), column=3).value}" #Locality Cell
            #cells are merged after for formatting and looks
            dest_sheet.merge_cells(start_row=(x+4), start_column=(i-2), end_row=(x+5), end_column=(i+1))
            dest_sheet.merge_cells(start_row=(x+6), start_column=(i - 2), end_row=(x + 7), end_column=(i + 1))
            count += 1

wb.save("CardOut.xlsx")
os.startfile("CardOut.xlsx")